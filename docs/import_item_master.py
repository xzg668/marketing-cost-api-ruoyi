#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_item_master.py
把 U9 料品主档 Excel 导入到 lp_material_master_raw 表。

特性：
  - 按 Excel 第 2 行表头映射到 DB 字段，不依赖固定列序
  - 支持模板版本 U9_ITEM_MASTER_20260519 和旧 U9 表头别名
  - 批量 executemany（默认每批 1000 行）
  - import_batch_id 默认按 u9-item-master-YYYYMMDD-HHMMSS 生成
  - 同一批次按物料代码去重，重复料号写到 failed.csv
  - 失败行写到 failed.csv，包含 Excel 行号、料号、失败原因

依赖：
  pip install openpyxl pymysql tqdm

用法：
  python3 docs/import_item_master.py \
      --excel /path/to/料品档案20260519.xlsx \
      --host 127.0.0.1 --port 3306 \
      --user root --db marketing_cost
"""
import argparse
import csv
import getpass
import sys
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parents[1] / "scripts"
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("缺少依赖：pip install openpyxl")
try:
    import pymysql
except ImportError:
    sys.exit("缺少依赖：pip install pymysql")
try:
    from tqdm import tqdm
except ImportError:
    class _NoopTqdm:
        def __init__(self, it, **kw):
            self.it = it

        def __iter__(self):
            return iter(self.it)

        def set_postfix(self, **kw):
            return None

        def close(self):
            return None

    def tqdm(it, **kw):
        return _NoopTqdm(it, **kw)

from u9_item_master_contract import DATA_START_ROW, MAPPING_VERSION, SHEET_NAME
from u9_item_master_import_mapping import (
    FIELD_NAMES,
    IMPORT_FIELD_NAMES,
    HeaderMappingError,
    MaterialCodeDeduplicator,
    build_header_field_index,
    build_import_row,
    describe_headers,
    material_code_from_import_row,
    read_header_values,
)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="/Users/xiexicheng/Documents/sales_cost/料品档案20260519.xlsx")
    ap.add_argument("--sheet", default=SHEET_NAME)
    ap.add_argument("--host", default="127.0.0.1")
    ap.add_argument("--port", type=int, default=3306)
    ap.add_argument("--user", default="root")
    ap.add_argument("--password", default=None, help="留空会交互询问")
    ap.add_argument("--db", default="marketing_cost")
    ap.add_argument("--batch-size", type=int, default=1000, help="每批 INSERT 行数")
    ap.add_argument("--batch-id", default=None, help="导入批次号（默认：u9-item-master-时间戳）")
    ap.add_argument("--truncate", action="store_true", help="导入前清空整张目标表（慎用）")
    ap.add_argument("--clear-batch", action="store_true", help="导入前清掉同 batch-id 的旧记录")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        sys.exit(f"ERROR: Excel 不存在：{excel_path}")

    pwd = args.password
    if pwd is None:
        pwd = getpass.getpass(f"请输入 MySQL 密码（{args.user}@{args.host}）: ")

    batch_id = args.batch_id or f"u9-item-master-{datetime.now():%Y%m%d-%H%M%S}"

    print("\n=========== 导入任务 ===========")
    print(f"  Excel       : {excel_path} ({excel_path.stat().st_size/1024/1024:.1f} MB)")
    print(f"  Sheet       : {args.sheet}")
    print(f"  目标库       : {args.db}@{args.host}:{args.port}")
    print("  目标表       : lp_material_master_raw")
    print(f"  batch_id    : {batch_id}")
    print(f"  mapping     : {MAPPING_VERSION}")
    print(f"  batch-size  : {args.batch_size}")
    if args.truncate:
        print("  TRUNCATE    : 整表清空")
    if args.clear_batch:
        print("  CLEAR       : 清掉同 batch-id 旧记录")
    print("================================\n")

    conn = pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=pwd,
        database=args.db,
        charset="utf8mb4",
        local_infile=False,
        autocommit=False,
        connect_timeout=30,
    )
    cur = conn.cursor()
    print("MySQL 连接 OK")

    cur.execute("SET autocommit = 0")
    cur.execute("SET unique_checks = 0")
    cur.execute("SET foreign_key_checks = 0")

    if args.truncate:
        cur.execute("TRUNCATE TABLE lp_material_master_raw")
        print("已清空目标表")
    elif args.clear_batch:
        cur.execute("DELETE FROM lp_material_master_raw WHERE import_batch_id=%s", (batch_id,))
        print(f"已清掉 batch_id={batch_id} 的旧记录：{cur.rowcount} 行")
    conn.commit()

    print("打开 Excel ...")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        sys.exit(f"ERROR: Sheet '{args.sheet}' 不存在，可用：{wb.sheetnames}")
    ws = wb[args.sheet]

    try:
        # 字段映射按表头匹配，避免 U9 导出列顺序变化导致默认供应商等字段错位。
        header_index = build_header_field_index(read_header_values(ws))
    except HeaderMappingError as exc:
        wb.close()
        sys.exit(f"ERROR: {exc}")

    header_desc = describe_headers(header_index)
    print(f"已映射表头: {header_desc['mapped']} / {len(FIELD_NAMES)}")
    if header_desc["missing_optional"]:
        print(f"可选表头缺失: {', '.join(header_desc['missing_optional'])}")

    cols = list(IMPORT_FIELD_NAMES)
    cols_sql = ",".join(f"`{col}`" for col in cols)
    placeholders = ",".join(["%s"] * len(cols))
    insert_sql = (
        f"INSERT INTO lp_material_master_raw ({cols_sql}) VALUES ({placeholders}) "
        "ON DUPLICATE KEY UPDATE imported_at = CURRENT_TIMESTAMP"
    )

    failed_csv = excel_path.with_suffix(".failed.csv")
    failed_f = open(failed_csv, "w", newline="", encoding="utf-8")
    failed_writer = csv.writer(failed_f)
    failed_writer.writerow(["excel_row", "material_code", "reason"])

    inserted = 0
    skipped = 0
    duplicated = 0
    failed = 0
    batch = []
    deduplicator = MaterialCodeDeduplicator()

    row_iter = ws.iter_rows(min_row=DATA_START_ROW, values_only=True)
    pbar = tqdm(row_iter, total=max(0, ws.max_row - DATA_START_ROW + 1), unit="行", desc="导入中")
    for excel_row, row in enumerate(pbar, start=DATA_START_ROW):
        values = build_import_row(row, header_index, batch_id)
        if values is None:
            skipped += 1
            continue
        material_code = material_code_from_import_row(values)
        first_row = deduplicator.accept(material_code, excel_row)
        if first_row is not None:
            duplicated += 1
            failed_writer.writerow([excel_row, material_code, f"重复物料代码，已使用第 {first_row} 行"])
            continue
        batch.append((excel_row, values))

        if len(batch) >= args.batch_size:
            ok, bad = flush_batch(cur, insert_sql, batch, failed_writer)
            inserted += ok
            failed += bad
            conn.commit()
            batch.clear()
            pbar.set_postfix(inserted=inserted, failed=failed, skipped=skipped, duplicated=duplicated)

    if batch:
        ok, bad = flush_batch(cur, insert_sql, batch, failed_writer)
        inserted += ok
        failed += bad
        conn.commit()
        batch.clear()

    pbar.close()

    cur.execute("SET unique_checks = 1")
    cur.execute("SET foreign_key_checks = 1")
    conn.commit()

    failed_f.close()
    wb.close()

    cur.execute("SELECT COUNT(*) FROM lp_material_master_raw WHERE import_batch_id=%s", (batch_id,))
    db_count = cur.fetchone()[0]

    print("\n========== 导入完成 ==========")
    print(f"  尝试插入 : {inserted + failed} 行")
    print(f"  成功插入 : {inserted} 行")
    print(f"  失败行   : {failed} 行（详见 {failed_csv}）")
    print(f"  跳过空料号 : {skipped} 行")
    print(f"  跳过重复料号 : {duplicated} 行")
    print(f"  当前批次 DB 行数 : {db_count}")
    print(f"  batch_id : {batch_id}")
    print("==============================")

    cur.close()
    conn.close()


def flush_batch(cur, sql, batch, failed_writer):
    rows = [item[1] for item in batch]
    try:
        cur.executemany(sql, rows)
        return len(rows), 0
    except Exception:
        ok = 0
        bad = 0
        for excel_row, row in batch:
            try:
                cur.execute(sql, row)
                ok += 1
            except Exception as exc:
                bad += 1
                failed_writer.writerow([excel_row, material_code_from_import_row(row), str(exc)[:300]])
        return ok, bad


if __name__ == "__main__":
    main()
