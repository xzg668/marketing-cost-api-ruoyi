import sys
import unittest
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = REPO_ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from u9_item_master_contract import DATA_START_ROW, FIELD_MAPPINGS
from u9_item_master_import_mapping import (
    FIELD_NAMES,
    HeaderMappingError,
    MaterialCodeDeduplicator,
    build_header_field_index,
    build_import_row,
    material_code_from_import_row,
    normalize_cell,
    read_header_values,
)


class U9ItemMasterImportMappingTest(unittest.TestCase):
    def test_maps_by_header_not_column_position(self):
        headers = [
            "默认主供应商",
            "物料代码*",
            "物料名称*",
            "采购处理提前期",
            "主分类代码",
            "主分类名称",
            "成本要素",
            "U9物料形态属性",
        ]
        row = ["供应商A", "MAT-001", "料品A", "7", "1101", "裸品", "材料", "采购件"]

        header_index = build_header_field_index(headers)
        values = build_import_row(row, header_index, "batch-001")
        by_field = dict(zip(list(FIELD_NAMES) + [
            "import_batch_id",
            "source_type",
            "source_batch_no",
            "mapping_version",
            "active_flag",
        ], values))

        self.assertEqual(by_field["default_supplier"], "供应商A")
        self.assertEqual(by_field["purchase_lead_time"], "7")
        self.assertEqual(by_field["main_category_code"], "1101")
        self.assertEqual(by_field["source_type"], "EXCEL")
        self.assertEqual(by_field["source_batch_no"], "batch-001")
        self.assertEqual(by_field["mapping_version"], "U9_ITEM_MASTER_20260519")
        self.assertEqual(by_field["active_flag"], 1)

    def test_old_header_aliases_are_supported(self):
        headers = [
            "物料代码*",
            "物料名称*",
            "料品采购相关信息.收货原则",
            "料品MRP相关信息.采购预处理提前期(天)",
            "料品采购相关信息.描述性弹性域.全局段3(理论净重)",
        ]
        row = ["MAT-002", "料品B", "按订单收货", "3", "47.8"]

        header_index = build_header_field_index(headers)
        values = build_import_row(row, header_index, "batch-002")
        by_field = dict(zip(FIELD_NAMES, values[:len(FIELD_NAMES)]))

        self.assertEqual(by_field["purchase_receive_principle"], "按订单收货")
        self.assertEqual(by_field["mrp_purchase_pre_lead_time"], "3")
        self.assertEqual(by_field["global_seg_3_theoretical_net_weight"], "47.8")

    def test_missing_required_header_blocks_import(self):
        with self.assertRaisesRegex(HeaderMappingError, "缺少必填表头: 物料代码"):
            build_header_field_index(["物料名称*", "默认主供应商"])

    def test_deduplicates_by_material_code_only(self):
        headers = ["裸品编码", "物料代码*", "物料名称*"]
        header_index = build_header_field_index(headers)
        rows = [
            (3, ["BARE-001", "MAT-003", "料品C"]),
            (4, ["BARE-002", "MAT-003", "料品C-重复"]),
        ]
        deduplicator = MaterialCodeDeduplicator()

        first_values = build_import_row(rows[0][1], header_index, "batch-003")
        second_values = build_import_row(rows[1][1], header_index, "batch-003")

        self.assertIsNone(deduplicator.accept(material_code_from_import_row(first_values), rows[0][0]))
        self.assertEqual(deduplicator.accept(material_code_from_import_row(second_values), rows[1][0]), 3)

    def test_new_contract_contains_63_fields(self):
        self.assertEqual(len(FIELD_MAPPINGS), 63)
        self.assertIn("global_seg_3_theoretical_net_weight", FIELD_NAMES)


class U9ItemMasterRealWorkbookTest(unittest.TestCase):
    def test_new_workbook_first_rows_do_not_shift_default_supplier(self):
        xlsx_path = Path("/Users/xiexicheng/Documents/sales_cost/料品档案20260519.xlsx")
        if not xlsx_path.exists():
            self.skipTest(f"Excel 不存在: {xlsx_path}")

        from openpyxl import load_workbook

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            sheet = workbook["物料主档"]
            header_index = build_header_field_index(read_header_values(sheet))
            self.assertEqual(header_index["default_supplier"], 18)
            self.assertEqual(header_index["purchase_lead_time"], 19)

            checked = 0
            differs_from_old_position = 0
            for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
                values = build_import_row(row, header_index, "batch-real")
                if values is None:
                    continue
                by_field = dict(zip(FIELD_NAMES, values[:len(FIELD_NAMES)]))
                self.assertEqual(by_field["default_supplier"], normalize_cell(row[18]))
                old_column_value = normalize_cell(row[43]) if len(row) > 43 else None
                if by_field["default_supplier"] != old_column_value:
                    differs_from_old_position += 1
                checked += 1
                if checked >= 10:
                    break

            self.assertEqual(checked, 10)
            self.assertGreater(differs_from_old_position, 0)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
