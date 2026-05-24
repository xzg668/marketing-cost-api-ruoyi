#!/usr/bin/env python3
"""
物料主档 (U9 ItemMaster Excel) 批量导入到 lp_material_master_raw。

用法：
    python3 scripts/import_item_master.py /path/to/料品档案20260519.xlsx

特点：
- 按第 2 行表头映射字段，支持模板版本 U9_ITEM_MASTER_20260519
- 支持旧 U9 表头别名
- 流式读 Excel，1000 行一批 INSERT IGNORE + commit
- 失败行写入 <Excel 文件名>.failed.csv，包含 Excel 行号、料号、失败原因
"""
import csv
import datetime
import os
import sys
import time
import warnings

warnings.filterwarnings("ignore")

import openpyxl
import pymysql

from u9_item_master_contract import DATA_START_ROW, MAPPING_VERSION, SHEET_NAME
from u9_item_master_import_mapping import (
    FIELD_NAMES,
    IMPORT_FIELD_NAMES,
    HeaderMappingError,
    MaterialCodeDeduplicator,
    build_header_field_index,
    build_import_row,
    describe_headers,
    material_code_from_import_row,
    read_header_values,
)


# ───────── 配置 ─────────
DB_HOST = "localhost"
DB_PORT = 3306
DB_USER = "root"
DB_PASS = "root123"
DB_NAME = "marketing_cost"
BATCH_SIZE = 1000


def fail(message):
    print(f"ERROR: {message}", file=sys.stderr)
    sys.exit(1)


def open_sheet(xlsx_path):
    print("打开 Excel ...")
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        fail(f"Sheet '{SHEET_NAME}' 不存在，可用：{workbook.sheetnames}")
    return workbook, workbook[SHEET_NAME]


def build_insert_sql():
    cols = list(IMPORT_FIELD_NAMES)
    placeholders = ",".join(["%s"] * len(cols))
    cols_sql = ",".join(f"`{col}`" for col in cols)
    return f"INSERT IGNORE INTO lp_material_master_raw ({cols_sql}) VALUES ({placeholders})"


def main(xlsx_path):
    if not os.path.exists(xlsx_path):
        fail(f"文件不存在: {xlsx_path}")

    batch_id = f"u9-item-master-{datetime.datetime.now():%Y%m%d-%H%M%S}"
    failed_csv = f"{os.path.splitext(xlsx_path)[0]}.failed.csv"
    print(f"导入批次: {batch_id}")
    print(f"映射版本: {MAPPING_VERSION}")

    t0 = time.time()
    workbook, sheet = open_sheet(xlsx_path)
    print(f"  sheet: {sheet.title}, 总行数: {sheet.max_row}")

    try:
        # 业务导入必须按表头匹配，避免 U9 导出列顺序变化导致字段错位。
        header_index = build_header_field_index(read_header_values(sheet))
    except HeaderMappingError as exc:
        workbook.close()
        fail(str(exc))

    header_desc = describe_headers(header_index)
    print(f"  已映射表头: {header_desc['mapped']} / {len(FIELD_NAMES)}")
    if header_desc["missing_optional"]:
        print(f"  可选表头缺失: {', '.join(header_desc['missing_optional'])}")

    conn = pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASS,
        database=DB_NAME,
        charset="utf8mb4",
        autocommit=False,
    )
    cursor = conn.cursor()
    sql = build_insert_sql()

    total_read = 0
    total_inserted = 0
    total_skipped = 0
    total_duplicated = 0
    total_failed = 0
    batch = []
    deduplicator = MaterialCodeDeduplicator()

    failed_file = open(failed_csv, "w", newline="", encoding="utf-8")
    failed_writer = csv.writer(failed_file)
    failed_writer.writerow(["excel_row", "material_code", "reason"])

    def flush(batch_data):
        nonlocal total_inserted, total_failed
        if not batch_data:
            return
        rows = [item[1] for item in batch_data]
        try:
            total_inserted += cursor.executemany(sql, rows)
            conn.commit()
            return
        except Exception:
            conn.rollback()

        for excel_row, values in batch_data:
            try:
                total_inserted += cursor.execute(sql, values)
            except Exception as exc:
                total_failed += 1
                material_code = material_code_from_import_row(values)
                failed_writer.writerow([excel_row, material_code, str(exc)[:300]])
        conn.commit()

    print("开始流式读取 + 批量入库 ...")
    for excel_row, row in enumerate(sheet.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        total_read += 1
        values = build_import_row(row, header_index, batch_id)
        if values is None:
            total_skipped += 1
            continue

        material_code = material_code_from_import_row(values)
        first_row = deduplicator.accept(material_code, excel_row)
        if first_row is not None:
            total_duplicated += 1
            failed_writer.writerow([excel_row, material_code, f"重复物料代码，已使用第 {first_row} 行"])
            continue

        batch.append((excel_row, values))
        if len(batch) >= BATCH_SIZE:
            flush(batch)
            batch = []
            if total_read % 10000 == 0:
                print(
                    f"  进度: 已读 {total_read} / 已入库 {total_inserted} "
                    f"/ 空料号 {total_skipped} / 重复料号 {total_duplicated} / 失败 {total_failed}"
                )

    flush(batch)

    failed_file.close()
    cursor.close()
    conn.close()
    workbook.close()

    elapsed = time.time() - t0
    print(f"\n导入完成 (耗时 {elapsed:.1f}s)")
    print(f"  总读取行数:   {total_read}")
    print(f"  成功入库:     {total_inserted}")
    print(f"  跳过(空料号): {total_skipped}")
    print(f"  跳过(重复料号): {total_duplicated}")
    print(f"  失败行数:     {total_failed}")
    print(f"  失败明细:     {failed_csv}")
    print(f"  批次 ID:      {batch_id}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("用法: python3 scripts/import_item_master.py <ItemMaster.xlsx>")
        sys.exit(1)
    main(sys.argv[1])
