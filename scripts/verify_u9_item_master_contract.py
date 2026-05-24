#!/usr/bin/env python3
"""校验 U9 料品主档字段契约和本次两份 Excel 的料号覆盖关系。"""
import argparse
import sys
import warnings
from pathlib import Path

from openpyxl import load_workbook

from u9_item_master_contract import (
    DATA_START_ROW,
    EXPECTED_HEADER_COUNT,
    EXPECTED_NEW_UNIQUE_CODES,
    EXPECTED_OLD_UNIQUE_CODES,
    FIELD_MAPPINGS,
    HEADER_ROW,
    MAPPING_VERSION,
    SHEET_NAME,
    canonical_header,
    resolve_field_by_header,
)


DEFAULT_NEW_XLSX = Path("/Users/xiexicheng/Documents/sales_cost/料品档案20260519.xlsx")
DEFAULT_OLD_XLSX = Path("/Users/xiexicheng/Desktop/ItemMaster20260427.xlsx")


def cell_to_text(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def open_material_sheet(path):
    if not path.exists():
        raise AssertionError(f"Excel 文件不存在: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise AssertionError(f"{path} 缺少工作表: {SHEET_NAME}")
    return workbook, workbook[SHEET_NAME]


def read_headers(sheet):
    row = next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    return [canonical_header(value) for value in row]


def header_positions(headers):
    positions = {}
    for index, header in enumerate(headers, start=1):
        if not header:
            continue
        if header in positions:
            raise AssertionError(f"表头重复: {header}")
        positions[header] = index
    return positions


def assert_new_headers(headers):
    named_headers = [header for header in headers if header]
    if len(named_headers) != EXPECTED_HEADER_COUNT:
        raise AssertionError(f"新 Excel 命名表头数量应为 63，实际为 {len(named_headers)}")
    if len(FIELD_MAPPINGS) != EXPECTED_HEADER_COUNT:
        raise AssertionError(f"字段契约数量应为 63，实际为 {len(FIELD_MAPPINGS)}")

    positions = header_positions(headers)
    for mapping in FIELD_MAPPINGS:
        actual_field = resolve_field_by_header(mapping.header)
        if actual_field != mapping.field:
            raise AssertionError(f"字段契约错误: {mapping.header} -> {actual_field}, 期望 {mapping.field}")
        actual_column = positions.get(mapping.header)
        if actual_column != mapping.excel_column:
            raise AssertionError(
                f"表头列号变化: {mapping.header} 实际第 {actual_column} 列，期望第 {mapping.excel_column} 列"
            )

    # 这几个字段是本次错位风险和新增字段的核心哨兵，后续导入必须按表头命中。
    sentinel_headers = ("物料代码*", "默认主供应商", "全局段3(理论净重)")
    for header in sentinel_headers:
        if header not in positions:
            raise AssertionError(f"关键表头未命中: {header}")


def collect_material_codes(path):
    workbook, sheet = open_material_sheet(path)
    try:
        positions = header_positions(read_headers(sheet))
        material_column = positions.get("物料代码*")
        if material_column is None:
            raise AssertionError(f"{path} 缺少表头: 物料代码*")

        codes = set()
        duplicates = set()
        data_rows = 0
        for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
            data_rows += 1
            value = row[material_column - 1] if len(row) >= material_column else None
            code = cell_to_text(value)
            if not code:
                continue
            if code in codes:
                duplicates.add(code)
            codes.add(code)
        return {
            "path": str(path),
            "data_rows": data_rows,
            "unique_codes": len(codes),
            "duplicates": duplicates,
            "codes": codes,
        }
    finally:
        workbook.close()


def verify(new_xlsx, old_xlsx):
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    workbook, sheet = open_material_sheet(new_xlsx)
    try:
        new_headers = read_headers(sheet)
        assert_new_headers(new_headers)
    finally:
        workbook.close()

    # 校验旧表头别名仍能被归并，避免兼容旧文件时把字段当成缺失。
    if resolve_field_by_header("料品采购相关信息.收货原则") != "purchase_receive_principle":
        raise AssertionError("旧表头别名未命中: 料品采购相关信息.收货原则")
    if resolve_field_by_header("料品MRP相关信息.采购预处理提前期(天)") != "mrp_purchase_pre_lead_time":
        raise AssertionError("旧表头别名未命中: 料品MRP相关信息.采购预处理提前期(天)")

    new_stats = collect_material_codes(new_xlsx)
    old_stats = collect_material_codes(old_xlsx)

    if new_stats["duplicates"]:
        raise AssertionError(f"新 Excel 存在重复料号: {len(new_stats['duplicates'])}")
    if old_stats["duplicates"]:
        raise AssertionError(f"原 Excel 存在重复料号: {len(old_stats['duplicates'])}")
    if new_stats["unique_codes"] != EXPECTED_NEW_UNIQUE_CODES:
        raise AssertionError(f"新 Excel 唯一料号数应为 166931，实际为 {new_stats['unique_codes']}")
    if old_stats["unique_codes"] != EXPECTED_OLD_UNIQUE_CODES:
        raise AssertionError(f"原 Excel 唯一料号数应为 163094，实际为 {old_stats['unique_codes']}")

    missing_codes = old_stats["codes"] - new_stats["codes"]
    if missing_codes:
        preview = ", ".join(sorted(missing_codes)[:10])
        raise AssertionError(f"新 Excel 未覆盖原料号: {len(missing_codes)}，示例: {preview}")

    added_codes = new_stats["codes"] - old_stats["codes"]
    return {
        "mapping_version": MAPPING_VERSION,
        "new_data_rows": new_stats["data_rows"],
        "new_unique_codes": new_stats["unique_codes"],
        "old_data_rows": old_stats["data_rows"],
        "old_unique_codes": old_stats["unique_codes"],
        "covered_old_codes": old_stats["unique_codes"],
        "missing_old_codes": len(missing_codes),
        "added_codes": len(added_codes),
    }


def main():
    parser = argparse.ArgumentParser(description="校验 U9 料品主档字段契约")
    parser.add_argument("--new-xlsx", type=Path, default=DEFAULT_NEW_XLSX)
    parser.add_argument("--old-xlsx", type=Path, default=DEFAULT_OLD_XLSX)
    args = parser.parse_args()

    result = verify(args.new_xlsx, args.old_xlsx)
    print(f"mapping_version: {result['mapping_version']}")
    print(f"new_data_rows: {result['new_data_rows']}")
    print(f"new_unique_codes: {result['new_unique_codes']}")
    print(f"old_data_rows: {result['old_data_rows']}")
    print(f"old_unique_codes: {result['old_unique_codes']}")
    print(f"covered_old_codes: {result['covered_old_codes']}")
    print(f"missing_old_codes: {result['missing_old_codes']}")
    print(f"added_codes: {result['added_codes']}")
    print("OK: U9MM-01 字段契约校验通过")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
